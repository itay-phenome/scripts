#!/usr/bin/env python3
"""
RDS Weekly Backup Checker — Phenome Networks
Scans S3 buckets for weekly RDS dump backups and writes three reports:
  <stem>_YYYY-MM-DD.json   — machine-readable
  <stem>_YYYY-MM-DD.html   — self-contained dashboard
  <stem>_YYYY-MM-DD.csv    — spreadsheet-friendly

Usage:
    python backup_checker.py
    python backup_checker.py --config backup_config.xlsx
    python backup_checker.py --config backup_config.xlsx --profile phenome-prod
    python backup_checker.py --config backup_config.xlsx --min-size-mb 200 --max-age-days 10
    python backup_checker.py --config backup_config.xlsx --output report_W19
    python backup_checker.py --dry-run          # validate config + credentials only
"""

import json, argparse, sys, os, re, time, csv, io
from datetime import datetime, timezone
from pathlib import Path

# ── constants ─────────────────────────────────────────────────────────────────
DATE_FORMAT        = "%Y-%m-%d"
DEFAULT_CONFIG     = Path(__file__).parent / "backup_config.xlsx"
DEFAULT_OUTPUT_STEM = Path(__file__).parent / "backup_report"
DEFAULT_MIN_MB     = 20
DEFAULT_MAX_DAYS   = 10
VALID_EXTENSIONS   = {".sql", ".gz", ".dump", ".tar", ".xb", ".qp", ".bak", ".zip"}
RETRY_CODES        = {"SlowDown", "RequestThrottled", "ServiceUnavailable", "InternalError"}
MAX_RETRIES        = 3
RETRY_DELAY_SEC    = 5
BUCKET_NAME_RE     = re.compile(r'^[a-z0-9][a-z0-9\-\.]{1,61}[a-z0-9]$')

# ── ANSI colour helpers ───────────────────────────────────────────────────────
USE_COLOR = sys.stdout.isatty() and os.name != "nt"
def _c(code): return f"\033[{code}m" if USE_COLOR else ""
R  = _c("0");  B  = _c("1");  GR = _c("32")
YL = _c("33"); RD = _c("31"); DM = _c("2")

def _ok(m):   print(f"  {GR}✓{R}  {m}")
def _warn(m): print(f"  {YL}⚠{R}  {YL}{m}{R}")
def _fail(m): print(f"  {RD}✗{R}  {RD}{m}{R}")
def _info(m): print(f"  {DM}·{R}  {DM}{m}{R}")
def _sec(t):  print(f"\n{B}{'─'*62}{R}\n{B}  {t}{R}\n{'─'*62}")
def _bytes_human(n):
    for u in ("B","KB","MB","GB","TB"):
        if n < 1024: return f"{n:.1f} {u}"
        n /= 1024
    return f"{n:.1f} PB"
def _iso_week(dt):
    iso = dt.isocalendar()
    return f"W{iso[1]:02d} / {iso[0]}"


# ══════════════════════════════════════════════════════════════════════════════
# STEP 0 — Python & library preflight
# ══════════════════════════════════════════════════════════════════════════════
def validate_python():
    _sec("STEP 0 — Python & library preflight")
    errors = []
    if sys.version_info < (3, 9):
        _fail(f"Python 3.9+ required (current: {sys.version.split()[0]})")
        errors.append("python")
    else:
        _ok(f"Python {sys.version.split()[0]}")
    try:
        import boto3
        _ok(f"boto3 {boto3.__version__}")
    except ImportError:
        _fail("boto3 not installed — run: pip install boto3")
        errors.append("boto3")
    try:
        import openpyxl
        _ok(f"openpyxl {openpyxl.__version__}")
    except ImportError:
        _warn("openpyxl not installed — required for .xlsx config (pip install openpyxl)")
    if errors:
        print(f"\n{RD}  Preflight failed. Fix the above and re-run.{R}\n")
        sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Config file validation
# ══════════════════════════════════════════════════════════════════════════════
def validate_config_file(config_path: Path) -> dict:
    _sec("STEP 1 — Config file validation")

    if not config_path.exists():
        _fail(f"Config file not found: {config_path}")
        sys.exit(1)
    _ok(f"File exists: {config_path.name}")

    if not os.access(config_path, os.R_OK):
        _fail(f"Config file is not readable: {config_path}")
        sys.exit(1)
    _ok("File is readable")

    suffix = config_path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        config = _parse_xlsx(config_path)
    elif suffix == ".json":
        config = _parse_json(config_path)
    else:
        _fail(f"Unsupported format: {suffix}  (use .xlsx or .json)")
        sys.exit(1)

    customers = config.get("customers", [])
    if not customers:
        _fail("No customers found in config file")
        sys.exit(1)
    _ok(f"{len(customers)} customers loaded")

    seen, warnings = {}, 0
    for i, c in enumerate(customers, 1):
        name   = str(c.get("name",   "")).strip()
        bucket = str(c.get("bucket", "")).strip()
        prefix = str(c.get("prefix", "")).strip()

        if not name:
            _warn(f"Row {i}: missing customer name — will be skipped")
            c["_skip"] = True; warnings += 1; continue

        if not c.get("enabled", True):
            _info(f"[{name}] disabled → will be skipped")
            c["_disabled"] = True; continue

        if name in seen:
            _warn(f"[{name}] duplicate (also row {seen[name]}) — second entry skipped")
            c["_skip"] = True; warnings += 1; continue
        seen[name] = i

        if not bucket and not prefix:
            _info(f"[{name}] not configured → N/A")
            c["_na"] = True; continue

        if bool(bucket) != bool(prefix):
            _warn(f"[{name}] only one of bucket/prefix is set → N/A")
            c["_na"] = True; warnings += 1; continue

        if bucket.startswith("s3://"):
            _warn(f"[{name}] stripping 's3://' from bucket name")
            c["bucket"] = bucket[5:].split("/")[0]; warnings += 1

        if not BUCKET_NAME_RE.match(c.get("bucket", bucket)):
            _warn(f"[{name}] bucket name looks invalid: '{bucket}'")
            warnings += 1

        if prefix.startswith("/"):
            _warn(f"[{name}] stripping leading '/' from prefix")
            c["prefix"] = prefix.lstrip("/"); warnings += 1

    cfg  = sum(1 for c in customers if not c.get("_skip") and not c.get("_na") and not c.get("_disabled"))
    na   = sum(1 for c in customers if c.get("_na"))
    skip = sum(1 for c in customers if c.get("_skip"))
    dis  = sum(1 for c in customers if c.get("_disabled"))
    _ok(f"Config valid — {cfg} to scan, {na} N/A, {dis} disabled, {skip} skipped")
    if warnings: _warn(f"{warnings} warning(s) noted above")
    return config


def _parse_json(path):
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        _ok("JSON parsed successfully")
        return data
    except json.JSONDecodeError as e:
        _fail(f"JSON parse error: {e}"); sys.exit(1)
    except Exception as e:
        _fail(f"Failed to read JSON: {e}"); sys.exit(1)


def _parse_xlsx(path):
    try:
        import openpyxl
    except ImportError:
        _fail("openpyxl required for .xlsx — run: pip install openpyxl"); sys.exit(1)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        _fail(f"Failed to open Excel file: {e}"); sys.exit(1)
    if "Backup Config" not in wb.sheetnames:
        _fail(f"Sheet 'Backup Config' not found. Sheets: {wb.sheetnames}"); sys.exit(1)
    _ok("Excel opened — sheet 'Backup Config' found")
    ws = wb["Backup Config"]
    customers = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        cells = (list(row) + [None]*7)[:7]
        _, name, group, bucket, prefix, _, enabled = cells
        if not name: continue
        enabled_str = str(enabled).strip().lower() if enabled is not None else "yes"
        customers.append({
            "name":    str(name).strip(),
            "group":   str(group).strip()  if group  else "",
            "bucket":  str(bucket).strip() if bucket else "",
            "prefix":  str(prefix).strip() if prefix else "",
            "enabled": enabled_str not in ("no", "false", "0", "disabled", "off"),
        })
    _ok(f"Read {len(customers)} rows from Excel")
    return {"reviewer": "Itay", "customers": customers}


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Output path validation
# ══════════════════════════════════════════════════════════════════════════════
def validate_output_paths(out_paths: list):
    _sec("STEP 2 — Output path validation")
    out_dir = out_paths[0].parent
    if not out_dir.exists():
        try:
            out_dir.mkdir(parents=True, exist_ok=True)
            _ok(f"Created output directory: {out_dir}")
        except Exception as e:
            _fail(f"Cannot create output directory {out_dir}: {e}"); sys.exit(1)
    else:
        _ok(f"Output directory exists: {out_dir}")

    test = out_dir / ".write_test"
    try:
        test.write_text("ok"); test.unlink()
        _ok(f"Output directory is writable: {out_dir}")
    except Exception as e:
        _fail(f"Cannot write to {out_dir}: {e}"); sys.exit(1)

    for p in out_paths:
        if p.exists():
            _warn(f"Will overwrite existing file: {p.name}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — AWS credentials validation
# ══════════════════════════════════════════════════════════════════════════════
def validate_aws_credentials(profile, region):
    _sec("STEP 3 — AWS credentials validation")
    try:
        import boto3
        import botocore
        from botocore.exceptions import NoCredentialsError, ClientError
    except ImportError:
        _fail("boto3 not available"); sys.exit(1)

    # profile exists?
    if profile:
        try:
            available = botocore.session.get_session().available_profiles
            if profile not in available:
                _fail(f"AWS profile '{profile}' not found. Available: {available}")
                sys.exit(1)
            _ok(f"AWS profile '{profile}' found")
        except Exception as e:
            _warn(f"Could not verify profile list: {e}")

    # STS identity check (avoids requiring s3:ListAllMyBuckets)
    try:
        session  = boto3.Session(profile_name=profile, region_name=region)
        sts      = session.client("sts")
        identity = sts.get_caller_identity()
        _ok(f"Credentials valid — Account: {identity['Account']}")
        _ok(f"Identity ARN: {identity['Arn']}")
    except NoCredentialsError:
        _fail("No AWS credentials found.")
        _info("Run: aws configure --profile <name>")
        _info("Or set: AWS_ACCESS_KEY_ID / AWS_SECRET_ACCESS_KEY")
        sys.exit(1)
    except Exception as e:
        _fail(f"Credential check failed: {e}"); sys.exit(1)

    s3 = session.client("s3")
    _ok("S3 client created successfully")
    return s3


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4a — Per-bucket pre-validation
# ══════════════════════════════════════════════════════════════════════════════
def validate_bucket(s3, bucket: str, prefix: str):
    """Returns (ok, error_note, regional_s3_client)."""
    import boto3
    from botocore.exceptions import ClientError

    # bucket accessible?
    try:
        s3.head_bucket(Bucket=bucket)
    except ClientError as e:
        code = e.response["Error"]["Code"]
        if code in ("404", "NoSuchBucket"):
            return False, f"Bucket does not exist: {bucket}", None
        if code in ("403", "AccessDenied"):
            return False, f"Access denied on bucket '{bucket}' — check IAM policy", None
        if code != "301":   # 301 = wrong region, handled below
            return False, f"HeadBucket error ({code})", None

    # resolve bucket region to avoid 301 redirect errors
    try:
        loc   = s3.get_bucket_location(Bucket=bucket)
        bregion = loc.get("LocationConstraint") or "us-east-1"
        regional_s3 = boto3.Session().client("s3", region_name=bregion)
    except ClientError as e:
        code = e.response["Error"]["Code"]
        if code in ("403", "AccessDenied"):
            regional_s3 = s3   # no GetBucketLocation permission — try with default
        else:
            return False, f"Could not get bucket region: {e}", None

    # prefix accessible + non-empty?
    try:
        resp = regional_s3.list_objects_v2(
            Bucket=bucket, Prefix=prefix.rstrip("/") + "/", MaxKeys=1
        )
        if resp.get("KeyCount", 0) == 0:
            return False, f"Prefix '{prefix}/' exists but is empty — backup job may not have run", regional_s3
    except ClientError as e:
        code = e.response["Error"]["Code"]
        if code in ("403", "AccessDenied"):
            return False, f"Access denied listing prefix '{prefix}/' — missing s3:ListBucket?", None
        return False, f"Error listing prefix: {e}", None

    return True, "", regional_s3


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4b — Per-backup content check
# ══════════════════════════════════════════════════════════════════════════════
def _s3_retry(fn):
    from botocore.exceptions import ClientError
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return fn()
        except ClientError as e:
            code = e.response["Error"]["Code"]
            if code in RETRY_CODES and attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY_SEC * attempt)
            else:
                raise


def check_backup(s3, bucket: str, prefix: str, min_bytes: int, max_age_days: int) -> dict:
    from botocore.exceptions import ClientError
    result = {"status": "missing", "size": "", "size_bytes": 0,
              "backupDate": "", "file_count": 0, "path": "", "notes": ""}

    prefix_norm = prefix.rstrip("/") + "/"
    pager       = s3.get_paginator("list_objects_v2")
    folders     = []

    # find date subfolders
    try:
        pages = _s3_retry(lambda: list(pager.paginate(Bucket=bucket, Prefix=prefix_norm, Delimiter="/")))
        for page in pages:
            for cp in page.get("CommonPrefixes", []):
                name = cp["Prefix"].rstrip("/").split("/")[-1]
                try:
                    datetime.strptime(name, DATE_FORMAT)
                    folders.append(name)
                except ValueError:
                    pass
    except ClientError as e:
        result["notes"] = f"S3 list error: {e.response['Error']['Code']}"
        return result

    if not folders:
        result["notes"] = f"No YYYY-MM-DD subfolders found under s3://{bucket}/{prefix}/"
        return result

    latest      = sorted(folders)[-1]
    full_prefix = prefix_norm + latest + "/"
    result["path"]       = f"s3://{bucket}/{full_prefix}"
    result["backupDate"] = latest

    # age check
    backup_dt = datetime.strptime(latest, DATE_FORMAT).replace(tzinfo=timezone.utc)
    age_days  = (datetime.now(timezone.utc) - backup_dt).days
    if age_days > max_age_days:
        result["status"] = "suspect"
        result["notes"]  = f"Latest backup is {age_days} day(s) old (max allowed: {max_age_days})"

    # list files in date folder
    objects = []
    try:
        pages = _s3_retry(lambda: list(pager.paginate(Bucket=bucket, Prefix=full_prefix)))
        for page in pages:
            objects.extend(page.get("Contents", []))
    except ClientError as e:
        result["notes"] = f"S3 list error inside date folder: {e.response['Error']['Code']}"
        return result

    files = [o for o in objects if o["Size"] > 0]
    result["file_count"] = len(files)

    if not files:
        result["status"] = "missing"
        result["notes"]  = "Date folder exists but contains no files"
        return result

    # file extension validation
    valid_files = [
        o for o in files
        if Path(o["Key"]).suffix.lower() in VALID_EXTENSIONS
        or any(o["Key"].endswith(e) for e in (".sql.gz", ".tar.gz", ".sql.bz2"))
    ]
    if not valid_files:
        exts = list({Path(o["Key"]).suffix for o in files})
        result["status"] = "suspect"
        result["notes"]  = (result["notes"] + "  " if result["notes"] else "") + \
                           f"No recognized backup file extensions (found: {exts})"
        return result

    invalid = [o for o in files if o not in valid_files]
    if invalid:
        exts = list({Path(o["Key"]).suffix for o in invalid})
        result["notes"] = (result["notes"] + "  " if result["notes"] else "") + \
                          f"{len(invalid)} unrecognized file(s) with ext(s): {exts}"

    # size check
    total = sum(o["Size"] for o in valid_files)
    result["size_bytes"] = total
    result["size"]       = _bytes_human(total)

    if total < min_bytes:
        result["status"] = "suspect"
        result["notes"]  = (result["notes"] + "  " if result["notes"] else "") + \
                           f"Size {_bytes_human(total)} below threshold ({_bytes_human(min_bytes)})"
        return result

    if result["status"] != "suspect":
        result["status"] = "ok"
    return result


# ══════════════════════════════════════════════════════════════════════════════
# Report writers
# ══════════════════════════════════════════════════════════════════════════════
def write_json_report(report: dict, path: Path):
    path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    json.loads(path.read_text(encoding="utf-8"))   # verify round-trip


def write_csv_report(report: dict, path: Path):
    fields = ["name", "group", "status", "backupDate", "size", "notes", "path"]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fields, extrasaction="ignore", lineterminator="\r\n")
    writer.writeheader()
    writer.writerows(report["rows"])
    # utf-8-sig BOM so Excel opens without a "convert" dialog
    path.write_text(buf.getvalue(), encoding="utf-8-sig")


def write_html_report(report: dict, path: Path):
    STATUS_COLOR = {"ok": "#22c55e", "missing": "#ef4444", "suspect": "#f59e0b", "na": "#9ca3af"}
    STATUS_BG    = {"ok": "#f0fdf4", "missing": "#fef2f2", "suspect": "#fffbeb", "na": "#f9fafb"}
    STATUS_LABEL = {"ok": "OK",      "missing": "MISSING", "suspect": "SUSPECT", "na": "N/A"}

    summary = report["summary"]
    flags   = report["flags"]

    rows_html = ""
    for r in report["rows"]:
        st    = r["status"]
        color = STATUS_COLOR.get(st, "#6b7280")
        bg    = STATUS_BG.get(st, "#ffffff")
        label = STATUS_LABEL.get(st, st.upper())
        rows_html += (
            f'<tr style="background:{bg}">'
            f'<td>{r["name"]}</td>'
            f'<td>{r.get("group","")}</td>'
            f'<td style="color:{color};font-weight:600">{label}</td>'
            f'<td>{r["backupDate"]}</td>'
            f'<td>{r["size"]}</td>'
            f'<td class="note">{r["notes"]}</td>'
            f'</tr>\n'
        )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Backup Report — {report["reportDate"]}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:system-ui,sans-serif;background:#f8fafc;color:#1e293b}}
  .hdr{{background:#1e293b;color:#fff;padding:1.25rem 2rem}}
  .hdr h1{{font-size:1.3rem;margin-bottom:.2rem}}
  .hdr .meta{{font-size:.8rem;opacity:.65}}
  .body{{padding:1.5rem 2rem}}
  .cards{{display:flex;gap:.75rem;margin-bottom:1.5rem;flex-wrap:wrap}}
  .card{{background:#fff;border-radius:8px;padding:.9rem 1.4rem;min-width:100px;
         box-shadow:0 1px 3px rgba(0,0,0,.08)}}
  .card .n{{font-size:1.9rem;font-weight:700;line-height:1}}
  .card .l{{font-size:.75rem;color:#64748b;margin-top:.2rem;text-transform:uppercase;letter-spacing:.04em}}
  .card.ok .n{{color:#22c55e}}.card.missing .n{{color:#ef4444}}
  .card.suspect .n{{color:#f59e0b}}.card.na .n{{color:#9ca3af}}
  table{{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;
         overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
  th{{background:#1e293b;color:#fff;padding:.55rem 1rem;text-align:left;
      font-size:.75rem;text-transform:uppercase;letter-spacing:.05em}}
  td{{padding:.5rem 1rem;border-bottom:1px solid #f1f5f9;font-size:.88rem}}
  tr:last-child td{{border-bottom:none}}
  td.note{{font-size:.8rem;color:#64748b}}
  .flags{{font-size:.78rem;color:#94a3b8;margin-top:1rem}}
</style>
</head>
<body>
<div class="hdr">
  <h1>RDS Backup Report — Phenome Networks</h1>
  <div class="meta">
    {report["reportWeek"]} &nbsp;·&nbsp; {report["reportDate"]}
    &nbsp;·&nbsp; Reviewer: {report["reviewer"]}
    &nbsp;·&nbsp; Generated: {report["generatedAt"]}
  </div>
</div>
<div class="body">
  <div class="cards">
    <div class="card ok">    <div class="n">{summary["ok"]}</div>      <div class="l">OK</div></div>
    <div class="card missing"><div class="n">{summary["missing"]}</div> <div class="l">Missing</div></div>
    <div class="card suspect"><div class="n">{summary["suspect"]}</div> <div class="l">Suspect</div></div>
    <div class="card na">    <div class="n">{summary["na"]}</div>       <div class="l">N/A</div></div>
    <div class="card">       <div class="n">{summary["total"]}</div>    <div class="l">Total</div></div>
  </div>
  <table>
    <thead><tr>
      <th>Customer</th><th>Group</th><th>Status</th>
      <th>Backup Date</th><th>Size</th><th>Notes</th>
    </tr></thead>
    <tbody>
{rows_html}    </tbody>
  </table>
  <div class="flags">
    Min size: {flags["minSizeMb"]} MB &nbsp;·&nbsp;
    Max age: {flags["maxAgeDays"]} days &nbsp;·&nbsp;
    AWS profile: {flags["awsProfile"]}
  </div>
</div>
</body>
</html>"""
    path.write_text(html, encoding="utf-8")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="RDS S3 Backup Checker — Phenome Networks")
    parser.add_argument("--config",       default=str(DEFAULT_CONFIG))
    parser.add_argument("--output",       default=str(DEFAULT_OUTPUT_STEM),
                        help="Output base name (no extension). Date is appended automatically. "
                             "Three files are written: <name>_YYYY-MM-DD.{json,html,csv}")
    parser.add_argument("--profile",      default=None)
    parser.add_argument("--region",       default="us-east-1")
    parser.add_argument("--min-size-mb",  type=int, default=DEFAULT_MIN_MB)
    parser.add_argument("--max-age-days", type=int, default=DEFAULT_MAX_DAYS)
    parser.add_argument("--dry-run",      action="store_true")
    args = parser.parse_args()

    if args.min_size_mb  <= 0: print(f"{RD}--min-size-mb must be > 0{R}",  file=sys.stderr); sys.exit(1)
    if args.max_age_days <= 0: print(f"{RD}--max-age-days must be > 0{R}", file=sys.stderr); sys.exit(1)

    config_path = Path(args.config)
    now         = datetime.now(timezone.utc)
    date_tag    = now.strftime(DATE_FORMAT)

    # Derive output paths — reports go into a date subfolder, e.g. .../2026-05-17/
    out_base = Path(args.output)
    out_dir  = out_base.parent / date_tag
    out_json = out_dir / f"{out_base.stem}_{date_tag}.json"
    out_html = out_dir / f"{out_base.stem}_{date_tag}.html"
    out_csv  = out_dir / f"{out_base.stem}_{date_tag}.csv"

    print(f"\n{B}{'═'*62}{R}")
    print(f"{B}  RDS Backup Checker — Phenome Networks{R}")
    print(f"{B}{'═'*62}{R}")
    _info(f"Config:        {config_path}")
    _info(f"Output (JSON): {out_json}")
    _info(f"Output (HTML): {out_html}")
    _info(f"Output (CSV):  {out_csv}")
    _info(f"Min size:      {args.min_size_mb} MB")
    _info(f"Max age:       {args.max_age_days} days")
    _info(f"AWS profile:   {args.profile or '(default)'}")
    if args.dry_run: _warn("DRY RUN — S3 scanning will be skipped")

    # ── pre-flight ────────────────────────────────────────────────────────────
    validate_python()
    config = validate_config_file(config_path)
    validate_output_paths([out_json, out_html, out_csv])
    s3 = validate_aws_credentials(args.profile, args.region)

    if args.dry_run:
        _sec("DRY RUN complete — all validations passed")
        customers = config.get("customers", [])
        to_scan = sum(1 for c in customers if not c.get("_skip") and not c.get("_na")
                      and c.get("bucket") and c.get("prefix"))
        print(f"\n{GR}  Ready to scan {to_scan} environments.{R}\n")
        sys.exit(0)

    # ── scan ──────────────────────────────────────────────────────────────────
    customers      = config.get("customers", [])
    min_bytes      = args.min_size_mb * 1024 * 1024

    _sec(f"STEP 4 — Scanning {len(customers)} environments")
    print(f"  {'#':<5} {'Customer':<38} {'Status':<18} {'Size':<10} {'Date'}")
    print(f"  {'─'*5} {'─'*38} {'─'*18} {'─'*10} {'─'*12}")

    rows = []
    for i, c in enumerate(customers, 1):
        name   = c.get("name",   f"Customer {i}")
        group  = c.get("group",  "")
        bucket = c.get("bucket", "").strip()
        prefix = c.get("prefix", "").strip().lstrip("/")

        if c.get("_skip"):
            print(f"  [{i:02d}]  {name:<38}  {DM}SKIPPED{R}")
            rows.append({"name": name, "group": group, "status": "na",
                         "size": "", "backupDate": "", "notes": "Skipped", "path": ""}); continue

        if c.get("_disabled"):
            print(f"  [{i:02d}]  {name:<38}  {DM}DISABLED{R}")
            rows.append({"name": name, "group": group, "status": "na",
                         "size": "", "backupDate": "", "notes": "Disabled", "path": ""}); continue

        if c.get("_na") or not bucket or not prefix:
            print(f"  [{i:02d}]  {name:<38}  {DM}N/A{R}")
            rows.append({"name": name, "group": group, "status": "na",
                         "size": "", "backupDate": "", "notes": "Not configured", "path": ""}); continue

        print(f"  [{i:02d}]  {name:<38}", end="  ", flush=True)

        ok, err, regional_s3 = validate_bucket(s3, bucket, prefix)
        if not ok:
            print(f"{RD}✗{R}  {RD}MISSING{R}")
            print(f"         {RD}└─ {err}{R}")
            rows.append({"name": name, "group": group, "status": "missing",
                         "size": "", "backupDate": "", "notes": err,
                         "path": f"s3://{bucket}/{prefix}/"}); continue

        result = check_backup(regional_s3, bucket, prefix, min_bytes, args.max_age_days)
        st     = result["status"]
        icons  = {"ok": f"{GR}✓{R}", "missing": f"{RD}✗{R}", "suspect": f"{YL}⚠{R}"}
        labels = {"ok": f"{GR}OK{R}", "missing": f"{RD}MISSING{R}", "suspect": f"{YL}SUSPECT{R}"}

        print(f"{icons.get(st,'?')}  {labels.get(st,st.upper()):<26}  {result['size']:<10}  {result['backupDate']}")
        if result["notes"]:
            nc = YL if st == "suspect" else RD
            print(f"         {nc}└─ {result['notes']}{R}")

        rows.append({"name": name, "group": group, "status": st,
                     "size": result["size"], "backupDate": result["backupDate"],
                     "notes": result["notes"], "path": result["path"]})

    # ── summary ───────────────────────────────────────────────────────────────
    ok_n  = sum(1 for r in rows if r["status"] == "ok")
    mis_n = sum(1 for r in rows if r["status"] == "missing")
    sus_n = sum(1 for r in rows if r["status"] == "suspect")
    na_n  = sum(1 for r in rows if r["status"] == "na")

    _sec("STEP 5 — Summary & output")
    print(f"  {GR}✓ OK:{R}       {ok_n}")
    print(f"  {RD}✗ Missing:{R}  {mis_n}")
    print(f"  {YL}⚠ Suspect:{R}  {sus_n}")
    print(f"  {DM}— N/A:{R}      {na_n}")

    if mis_n:
        print(f"\n  {RD}Missing environments:{R}")
        for r in rows:
            if r["status"] == "missing":
                print(f"    {RD}✗{R}  {r['name']}  —  {r['notes']}")
    if sus_n:
        print(f"\n  {YL}Suspect environments:{R}")
        for r in rows:
            if r["status"] == "suspect":
                print(f"    {YL}⚠{R}  {r['name']}  —  {r['notes']}")

    # ── build report dict ─────────────────────────────────────────────────────
    report = {
        "reportDate":  date_tag,
        "reportWeek":  _iso_week(now),
        "reviewer":    config.get("reviewer", ""),
        "generatedAt": now.isoformat(),
        "flags":       {"minSizeMb": args.min_size_mb, "maxAgeDays": args.max_age_days,
                        "awsProfile": args.profile or "(default)"},
        "summary":     {"total": len(rows), "ok": ok_n, "missing": mis_n, "suspect": sus_n, "na": na_n},
        "rows":        rows,
    }

    # ── write all three formats ───────────────────────────────────────────────
    try:
        write_json_report(report, out_json)
        _ok(f"JSON  → {out_json}")
    except Exception as e:
        _fail(f"Failed to write JSON: {e}"); sys.exit(1)

    try:
        write_html_report(report, out_html)
        _ok(f"HTML  → {out_html}")
    except Exception as e:
        _fail(f"Failed to write HTML: {e}"); sys.exit(1)

    try:
        write_csv_report(report, out_csv)
        _ok(f"CSV   → {out_csv}")
    except Exception as e:
        _fail(f"Failed to write CSV: {e}"); sys.exit(1)

    print()
    sys.exit(1 if mis_n > 0 else 0)   # non-zero exit if any missing (for cron/CI)


if __name__ == "__main__":
    main()
